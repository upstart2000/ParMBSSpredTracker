"""
Parses the FINRA-ICE Data Services Structured Product Pricing Tables (FINRA_IDS_PXTABLES.xlsx),
TBA sheet, 30-Year UMBS AVERAGE PRICE rows, and computes the par coupon via linear interpolation.
"""
import openpyxl
from datetime import date


TARGET_TABLE_LABEL = "PRICING TABLE: AGENCY PASS-THRU (TBA, STIP, $ ROLLS) - SINGLE FAMILY 30Y"
SUB_CLASS = "UMBS"
PRICE_METRIC_ROW = "AVERAGE PRICE"
COUPON_STEP = 0.5
DATA_AS_OF_LABEL = "DATA AS OF:"


def get_data_as_of_date(filepath):
    """
    Scans the TBA sheet's top banner for the "DATA AS OF:" label and returns the
    adjacent date cell as a date. This is FINRA's own stamp of which trading day
    the file's prices cover - used to detect whether a freshly-downloaded file has
    actually been refreshed yet (vs. still serving the prior day's data).
    Returns None if the label isn't found (unexpected sheet layout).
    """
    import datetime as dt

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["TBA"]

    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == DATA_AS_OF_LABEL:
                adjacent = ws.cell(row=cell.row, column=cell.column + 1).value
                if isinstance(adjacent, dt.datetime):
                    return adjacent.date()
                if isinstance(adjacent, dt.date):
                    return adjacent
                if isinstance(adjacent, str):
                    return dt.datetime.fromisoformat(adjacent.strip()).date()
    return None


def parse_tba_30y_umbs(filepath):
    """
    Returns dict: { settlement_month_label (e.g. 'August'): {coupon(float): price(float)} }
    Only numeric coupon columns are included (the '<= 3.5' and '> 6' edge buckets are dropped,
    since they can't serve as interpolation endpoints).
    Cells showing '*' (thin-trading suppression) or 0 (no trades) are dropped as unusable.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["TBA"]

    results = {}
    row = 1
    max_row = ws.max_row

    while row <= max_row:
        cell_b = ws.cell(row=row, column=2).value
        if isinstance(cell_b, str) and cell_b.strip() == TARGET_TABLE_LABEL:
            # Next row: "<Month> Settlement"
            settlement_row_val = ws.cell(row=row + 1, column=2).value
            if not isinstance(settlement_row_val, str) or "Settlement" not in settlement_row_val:
                row += 1
                continue
            month_label = settlement_row_val.replace("Settlement", "").strip()

            # Walk forward through sub-class blocks (UMBS, FHLMC, GNMA) until next
            # blank/table-label row, to find the UMBS coupon header + AVERAGE PRICE row.
            r = row + 3  # row+2 is "Asset Sub-Class / Metric | COUPON" header
            coupon_prices = {}
            while r <= max_row:
                sub_class_cell = ws.cell(row=r, column=2).value
                if sub_class_cell is None:
                    break  # blank row => end of this settlement block
                if isinstance(sub_class_cell, str) and sub_class_cell.strip() == TARGET_TABLE_LABEL:
                    break  # ran into next table

                if isinstance(sub_class_cell, str) and sub_class_cell.strip() == SUB_CLASS:
                    # This row is the coupon header row for UMBS: columns C.. = coupon labels
                    coupon_headers = [ws.cell(row=r, column=c).value for c in range(3, 10)]
                    # Find the AVERAGE PRICE row a few rows below
                    pr = r + 1
                    while pr <= max_row:
                        metric_cell = ws.cell(row=pr, column=2).value
                        if isinstance(metric_cell, str) and metric_cell.strip() == PRICE_METRIC_ROW:
                            price_vals = [ws.cell(row=pr, column=c).value for c in range(3, 10)]
                            for coup, px in zip(coupon_headers, price_vals):
                                if not isinstance(coup, (int, float)):
                                    continue  # skip '<= 3.5' / '> 6' text buckets
                                if not isinstance(px, (int, float)):
                                    continue  # skip '*' suppressed cells
                                if px == 0:
                                    continue  # skip no-trade cells
                                coupon_prices[float(coup)] = float(px)
                            break
                        pr += 1
                    r += 1
                    continue
                r += 1

            if coupon_prices:
                results[month_label] = coupon_prices
        row += 1

    return results


def compute_par_coupon(coupon_prices):
    """
    Linear interpolation between the coupon bucket just below par (price <= 100)
    and the one just above (price > 100), among usable (non-suppressed, non-zero) buckets.
    Returns None if no valid bracketing pair exists (e.g. all coupons priced above or below par).
    """
    if not coupon_prices:
        return None

    coupons_sorted = sorted(coupon_prices.keys())

    c_low = p_low = c_high = p_high = None
    for c in coupons_sorted:
        p = coupon_prices[c]
        if p <= 100:
            c_low, p_low = c, p  # keep updating to the highest-coupon sub-100 bucket
        elif p > 100 and c_high is None:
            c_high, p_high = c, p  # first coupon above par, immediately after the sub-100 run
            break

    if c_low is None or c_high is None:
        return None  # par coupon out of range of available brackets

    if p_high == p_low:
        return None  # degenerate, avoid divide-by-zero

    par_coupon = c_low + (100 - p_low) * (c_high - c_low) / (p_high - p_low)
    return round(par_coupon, 4), (c_low, p_low), (c_high, p_high)


if __name__ == "__main__":
    filepath = "/mnt/user-data/uploads/FINRA_IDS_PXTABLES.xlsx"
    parsed = parse_tba_30y_umbs(filepath)
    print("Settlement months found:", list(parsed.keys()))
    for month, cp in parsed.items():
        print(f"\n{month} Settlement, UMBS AVERAGE PRICE by coupon:")
        for c, p in sorted(cp.items()):
            print(f"  {c}: {p}")
        result = compute_par_coupon(cp)
        if result:
            par, (c_low, p_low), (c_high, p_high) = result
            print(f"  -> Par coupon = {par}%  (bracket: {c_low}@{p_low} / {c_high}@{p_high})")
        else:
            print("  -> Par coupon: not computable (no valid bracket)")
